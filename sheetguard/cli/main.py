# sheetguard/cli/main.py
from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import sys
import traceback
import webbrowser
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import networkx as nx
import yaml
from openpyxl import load_workbook

# openpyxl exception compatibility (varies by version)
try:
    from openpyxl.utils.exceptions import InvalidFileException  # type: ignore
except Exception:  # pragma: no cover
    InvalidFileException = Exception  # fallback

from sheetguard.core.graph import (
    build_dependency_graph,
    detect_cycles,
    fan_out,
)
from sheetguard.core.references import (
    RefToken,
    detect_volatile_functions,
    extract_reference_tokens,
)
from sheetguard.detectors.impact import rank_high_risk_cells
from sheetguard.detectors.constants import detect_hardcoded_constants
from sheetguard.detectors.drift import detect_formula_drift
from sheetguard.reporting.html_report import render_html_report


def _utc_iso() -> str:
    return _dt.datetime.utcnow().replace(tzinfo=_dt.timezone.utc).isoformat().replace("+00:00", "Z")


def _is_formula_cell(cell) -> bool:
    v = getattr(cell, "value", None)
    return isinstance(v, str) and v.startswith("=")


def _iter_cells(ws):
    """Yield populated cells.

    Note: ws._cells may exist but be empty until the worksheet has been iterated.
    """
    cells = getattr(ws, "_cells", None)
    if isinstance(cells, dict) and len(cells) > 0:
        for c in cells.values():
            yield c
        return
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            yield cell


def _load_rules_config() -> Dict[str, Any]:
    """
    Loads sheetguard/config/rules.yaml.

    Backward compatible: if config doesn't define app/trial, we use defaults.
    """
    cfg_path = Path(__file__).resolve().parents[1] / "config" / "rules.yaml"
    try:
        raw = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        if not isinstance(raw, dict):
            return {}
        return raw
    except Exception:
        return {}


def _cfg_get(cfg: Dict[str, Any], path: str, default: Any = None) -> Any:
    """Safely fetch a dotted-path value from nested dict configs."""
    if not path:
        return default
    cur: Any = cfg
    for key in str(path).split("."):
        if not isinstance(cur, dict) or key not in cur:
            return default
        cur = cur.get(key)
    return cur


def _trial_settings(cfg: Dict[str, Any], trial_flag: bool) -> Dict[str, Any]:
    """
    Trial is enabled either by CLI/UI flag OR by config default.
    """
    app = cfg.get("app", {}) if isinstance(cfg.get("app", {}), dict) else {}
    trial_cfg = app.get("trial", {}) if isinstance(app.get("trial", {}), dict) else {}

    enabled_default = bool(trial_cfg.get("enabled_default", False))
    limit = int(trial_cfg.get("formula_limit", 100) or 100)

    enabled = bool(trial_flag or enabled_default)
    return {"enabled": enabled, "formula_limit": max(1, limit)}


def _resolve_output_dir(workbook_path: str, out_dir: str) -> Path:
    """
    Easy output rule:
    - If out_dir is a directory (UI passes workbook parent), write to:
        <out_dir>/output/<workbook_stem>/<YYYYMMDD_HHMMSS>/
    - If out_dir looks like a file path or an explicit folder, we still create a unique run folder inside it.
    """
    wb = Path(workbook_path)
    base = Path(out_dir) if out_dir else wb.parent

    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = wb.stem

    # If user passed the workbook folder, keep things tidy under /output/<stem>/...
    # If user passed "output", also works.
    return (base / "output" / stem / ts).resolve()


def scan_workbook(
    workbook_path: str,
    out_dir: str,
    top_n: int = 10,
    trial: bool = False,
) -> Tuple[Dict[str, Any], str, str]:
    """
    Deterministic, read-only scan.
    Returns: (scan_dict, json_path, html_path)
    """
    p = Path(workbook_path)
    if not p.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    cfg = _load_rules_config()
    trial_info = _trial_settings(cfg, trial)

    outp = _resolve_output_dir(workbook_path=str(p), out_dir=out_dir)
    outp.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(filename=str(p), data_only=False, read_only=False, keep_vba=True)

    errors: List[Dict[str, Any]] = []
    g = nx.DiGraph()
    formula_nodes: List[str] = []
    symbolic_edges: List[Dict[str, Any]] = []
    # capture formula text for reporting (e.g., orphan formulas)
    formula_by_node: Dict[str, str] = {}
    formula_cells_by_sheet: Dict[str, List[Tuple[str, str]]] = {}

    # workbook-level counters
    volatile_hits = 0
    symbolic_refs = 0
    hardcoded_constants_hits = 0
    hardcoded_constants_findings: List[Dict[str, Any]] = []
    formula_scanned = 0
    truncated = False

    # per-sheet stats
    per_sheet_formula_counts: Dict[str, int] = {}
    per_sheet_signals: Dict[str, Dict[str, int]] = {}

    for ws in wb.worksheets:
        sheet = ws.title
        per_sheet_formula_counts[sheet] = 0
        per_sheet_signals[sheet] = {"volatile": 0, "symbolic": 0, "constants": 0, "orphans": 0, "drift": 0}
        formula_cells_by_sheet[sheet] = []

        try:
            for cell in _iter_cells(ws):
                if not _is_formula_cell(cell):
                    continue

                # Trial gating (formula count)
                if trial_info["enabled"] and formula_scanned >= int(trial_info["formula_limit"]):
                    truncated = True
                    break

                addr = getattr(cell, "coordinate", "")
                formula = str(getattr(cell, "value", "") or "")

                # keep for drift detector
                formula_cells_by_sheet[sheet].append((addr, formula))

                # keep formula text for downstream detectors/reporting
                formula_by_node[f"{sheet}!{addr}"] = formula
                formula_by_node[f"{sheet}!{addr}"] = formula

                per_sheet_formula_counts[sheet] += 1
                formula_scanned += 1

                # volatile hits
                try:
                    v_hits = detect_volatile_functions(formula)
                    volatile_hits += int(len(v_hits))
                    per_sheet_signals[sheet]["volatile"] += int(len(v_hits))
                except Exception as e:
                    errors.append(
                        {
                            "scope": "parse_formula",
                            "type": "volatile_detect_failed",
                            "dependent": f"{sheet}!{addr}",
                            "ref": "",
                            "details": repr(e),
                        }
                    )
                    v_hits = []

                # reference tokenization (RefToken objects)
                try:
                    raw_tokens = extract_reference_tokens(formula)
                    tokens = [RefToken.from_any(t) for t in (raw_tokens or [])]
                except Exception as e:
                    errors.append(
                        {
                            "scope": "parse_formula",
                            "type": "extract_reference_tokens_failed",
                            "dependent": f"{sheet}!{addr}",
                            "ref": "",
                            "details": repr(e),
                        }
                    )
                    tokens = []

                # count symbolic refs
                for t in tokens:
                    if bool(getattr(t, "is_symbolic", False)):
                        symbolic_refs += 1
                        per_sheet_signals[sheet]["symbolic"] += 1

                # hard-coded constants (config-driven)
                try:
                    const_cfg = _cfg_get(cfg, "app.detectors.hardcoded_constants", {}) or {}
                    if bool(const_cfg.get("enabled", True)):
                        hits = detect_hardcoded_constants(formula, cfg=const_cfg)
                    else:
                        hits = []
                    hardcoded_constants_hits += int(len(hits))
                    per_sheet_signals[sheet]["constants"] += int(len(hits))

                    # ensures we keep a bounded list for report readability
                    max_findings = int(const_cfg.get("max_findings", 100) or 100)
                    if max_findings > 0 and hits:
                        for h in hits:
                            if len(hardcoded_constants_findings) >= max_findings:
                                break
                            hardcoded_constants_findings.append(
                                h.to_dict(sheet=sheet, cell=addr, formula=formula)
                            )
                except Exception as e:
                    errors.append(
                        {
                            "scope": "parse_formula",
                            "type": "hardcoded_constants_detect_failed",
                            "dependent": f"{sheet}!{addr}",
                            "ref": "",
                            "details": repr(e),
                        }
                    )

                # graph edges (Jan-28 v1 incremental signature)
                try:
                    build_dependency_graph(
                        sheet_name=sheet,
                        cell_addr=addr,
                        refs=tokens,
                        g=g,
                        symbolic_edges=symbolic_edges,
                    )
                    formula_nodes.append(f"{sheet}!{addr}")
                except Exception as e:
                    errors.append(
                        {
                            "scope": "graph_build",
                            "type": "build_dependency_graph_failed",
                            "dependent": f"{sheet}!{addr}",
                            "ref": "",
                            "details": repr(e),
                        }
                    )

            if truncated:
                break

        except Exception as e:
            errors.append(
                {
                    "scope": "sheet_scan",
                    "type": "sheet_failed",
                    "dependent": sheet,
                    "ref": "",
                    "details": repr(e),
                }
            )

    # cycles
    try:
        cycles = detect_cycles(g, limit=25)
    except Exception as e:
        cycles = []
        errors.append({"scope": "cycles", "type": "detect_cycles_failed", "dependent": "", "ref": "", "details": repr(e)})

    # impact ranking
    try:
        top_risks = rank_high_risk_cells(g, formula_nodes=formula_nodes, top_n=int(top_n))
    except Exception as e:
        top_risks = []
        errors.append(
            {
                "scope": "impact",
                "type": "rank_top_impact_failed",
                "dependent": "",
                "ref": "",
                "details": repr(e),
            }
        )

    

    # Step 2C: copied formula drift (config-driven, conservative)
    drift_findings: List[Dict[str, Any]] = []
    drift_hits = 0
    try:
        drift_cfg = _cfg_get(cfg, "app.detectors.formula_drift", {}) or {}
        if bool(drift_cfg.get("enabled", True)):
            drift_findings = detect_formula_drift(formula_cells_by_sheet, drift_cfg) or []
            drift_hits = int(len(drift_findings))
            # attribute drift per sheet
            for it in drift_findings:
                sh = str(it.get("sheet", "") or "")
                if sh in per_sheet_signals:
                    per_sheet_signals[sh]["drift"] = int(per_sheet_signals[sh].get("drift", 0) or 0) + 1
    except Exception as e:
        errors.append(
            {
                "scope": "detector",
                "type": "formula_drift_failed",
                "dependent": "",
                "ref": "",
                "details": repr(e),
            }
        )
        drift_findings = []
        drift_hits = 0

# Step 2B: orphan calculations (config-driven)
    orphan_findings: List[Dict[str, Any]] = []
    orphan_hits = 0
    try:
        orphan_cfg = _cfg_get(cfg, "app.detectors.orphan_calculations", {}) or {}
        if bool(orphan_cfg.get("enabled", True)):
            exclude_sheets = set([str(s) for s in (orphan_cfg.get("exclude_sheets", []) or [])])
            exclude_named = bool(orphan_cfg.get("exclude_named_ranges", True))
            exclude_last = bool(orphan_cfg.get("exclude_last_row_or_col", True))
            max_findings = int(orphan_cfg.get("max_findings", 100) or 100)

            # Build a quick lookup: which nodes are formula nodes?
            formula_set = set(formula_nodes)

            # Named-range cell destinations (optional)
            named_cells: Set[str] = set()
            if exclude_named:
                try:
                    for dn in getattr(wb, "defined_names", []):
                        try:
                            for _title, coord in dn.destinations:
                                if coord:
                                    # coord can be 'A1' or 'A1:B2'
                                    part = str(coord).split(":", 1)[0]
                                    named_cells.add(f"{_title}!{part}")
                        except Exception:
                            continue
                except Exception:
                    named_cells = set()

            # Sheet max row/col (for excluding terminal row/col)
            sheet_max: Dict[str, Tuple[int, int]] = {}
            if exclude_last:
                for ws in wb.worksheets:
                    try:
                        sheet_max[ws.title] = (int(getattr(ws, "max_row", 0) or 0), int(getattr(ws, "max_column", 0) or 0))
                    except Exception:
                        sheet_max[ws.title] = (0, 0)

            # Orphan logic: no incoming edges from other formula nodes
            for node in list(formula_nodes):
                if "!" in str(node):
                    sh, addr = str(node).split("!", 1)
                else:
                    sh, addr = "", str(node)
                if not sh or not addr:
                    continue
                if sh in exclude_sheets:
                    continue
                if exclude_named and node in named_cells:
                    continue
                if exclude_last:
                    mr, mc = sheet_max.get(sh, (0, 0))
                    try:
                        cell_obj = wb[sh][addr]
                        if (mr and cell_obj.row == mr) or (mc and cell_obj.column == mc):
                            continue
                    except Exception:
                        pass

                preds = [p for p in g.predecessors(node) if p in formula_set]
                if len(preds) == 0:
                    orphan_hits += 1
                    per_sheet_signals.setdefault(sh, {}).setdefault("orphans", 0)
                    per_sheet_signals[sh]["orphans"] = int(per_sheet_signals[sh].get("orphans", 0) or 0) + 1

                    if max_findings > 0 and len(orphan_findings) < max_findings:
                        orphan_findings.append(
                            {
                                "type": "ORPHAN_FORMULA",
                                "sheet": sh,
                                "cell": addr,
                                "formula": formula_by_node.get(node, ""),
                                "reason": "formula not referenced by any other formula cell",
                            }
                        )
    except Exception as e:
        errors.append(
            {
                "scope": "orphans",
                "type": "orphan_detect_failed",
                "dependent": "",
                "ref": "",
                "details": repr(e),
            }
        )

    
    # per-sheet output blocks (reporting aggregation)
    # We derive sheet-level metrics from already-collected signals + graph findings.
    # No hardcoded thresholds: scoring is driven by config with safe defaults.
    def _split_cell_ref(s: str) -> Tuple[str, str]:
        if not s:
            return "", ""
        if "!" in s:
            sh, addr = s.split("!", 1)
            return sh, addr
        return "", s

    # circular cells per sheet
    circular_cells_by_sheet: Dict[str, set] = {sh: set() for sh in per_sheet_formula_counts.keys()}
    for c in (cycles or []):
        for ref in (c.get("cells", []) or []):
            sh, addr = _split_cell_ref(str(ref))
            if sh:
                circular_cells_by_sheet.setdefault(sh, set()).add(addr)

    # high-risk cells per sheet (Top-N impact list)
    top_by_sheet: Dict[str, int] = {sh: 0 for sh in per_sheet_formula_counts.keys()}
    for r in (top_risks or []):
        sh, _addr = _split_cell_ref(str(r.get("cell", "")))
        if sh:
            top_by_sheet[sh] = int(top_by_sheet.get(sh, 0) or 0) + 1

    # risk scoring (conservative defaults; can be overridden in config)
    scoring = _cfg_get(cfg, "app.reporting.risk_scoring", {}) or {}

    def _compute_risk_tier(
        *,
        scope: str,
        signals_map: Dict[str, int],
        has_cycles: bool,
        top_risk_cells: int,
        scoring_cfg: Dict[str, Any],
    ) -> Tuple[str, str]:
        """Return (tier, reason).

        Option A: do not expose numeric scores in the report.
        Scoring is config-driven when weights/thresholds exist; otherwise fall back to
        conservative boolean rules.
        """
        # --- Preferred: weighted scoring when config provides it ---
        thresholds = _cfg_get(scoring_cfg, f"{scope}.thresholds", None)
        weights = _cfg_get(scoring_cfg, f"{scope}.weights", None)
        overrides = _cfg_get(scoring_cfg, "overrides", {}) or {}

        # Override: any cycle can force HIGH
        if bool(overrides.get("cycle_always_high", True)) and bool(has_cycles):
            reason = "circular dependency detected"
            return "HIGH", reason

        if isinstance(thresholds, dict) and isinstance(weights, dict):
            # Normalize some signals to reduce scale issues (still deterministic)
            sym = int(signals_map.get("symbolic", 0) or 0)
            vol = int(signals_map.get("volatile", 0) or 0)
            consts = int(signals_map.get("constants", 0) or 0)
            # orphan not yet implemented in this branch; keep for forward compatibility
            orphans = int(signals_map.get("orphans", 0) or 0)

            # Weighted contributions
            contrib: Dict[str, float] = {}
            contrib["cycles"] = float(weights.get("cycles", 0) or 0) * (1.0 if has_cycles else 0.0)
            contrib["top_risk_cells"] = float(weights.get("top_risk_cells", 0) or 0) * float(top_risk_cells)
            contrib["volatile_hits"] = float(weights.get("volatile_hits", 0) or 0) * float(vol)
            contrib["hardcoded_constants"] = float(weights.get("hardcoded_constants", 0) or 0) * float(consts)
            contrib["orphan_formulas"] = float(weights.get("orphan_formulas", 0) or 0) * float(orphans)

            # symbolic is often large; score per 100 to keep it interpretable
            sym_per_100 = sym / 100.0
            contrib["symbolic_refs_per_100"] = float(weights.get("symbolic_refs_per_100", 0) or 0) * float(sym_per_100)

            score = sum(contrib.values())
            high_th = float(thresholds.get("high", 999999) or 999999)
            med_th = float(thresholds.get("medium", 999999) or 999999)

            if score >= high_th:
                tier = "HIGH"
            elif score >= med_th:
                tier = "MEDIUM"
            else:
                tier = "LOW"

            # Reason: top contributing non-zero factors (deterministic)
            labels = {
                "cycles": "circular dependency",
                "top_risk_cells": "high-impact drivers",
                "volatile_hits": "volatile functions",
                "hardcoded_constants": "embedded multipliers",
                "orphan_formulas": "orphan formulas",
                "symbolic_refs_per_100": "symbolic references",
            }
            ranked = sorted([(k, v) for k, v in contrib.items() if v > 0], key=lambda kv: kv[1], reverse=True)
            top_factors = [labels.get(k, k) for k, _v in ranked[:3]]
            reason = "; ".join(top_factors)
            return tier, reason

        # --- Fallback: conservative boolean rules (no numeric constants in code) ---
        # Sheet scoring toggles
        treat_symbolic_as_medium = bool(_cfg_get(scoring_cfg, f"{scope}.symbolic_counts_as_medium", True))
        treat_toprisk_as_medium = bool(_cfg_get(scoring_cfg, f"{scope}.top_risk_counts_as_medium", True))
        treat_volatile_as_medium = bool(_cfg_get(scoring_cfg, f"{scope}.volatile_counts_as_medium", True))
        treat_constants_as_medium = bool(_cfg_get(scoring_cfg, f"{scope}.constants_counts_as_medium", True))
        treat_orphans_as_medium = bool(_cfg_get(scoring_cfg, f"{scope}.orphans_counts_as_medium", True))
        cycles_make_high = bool(_cfg_get(scoring_cfg, f"{scope}.cycles_make_high", True))

        sym = int(signals_map.get("symbolic", 0) or 0)
        vol = int(signals_map.get("volatile", 0) or 0)
        consts = int(signals_map.get("constants", 0) or 0)
        orphans = int(signals_map.get("orphans", 0) or 0)

        if cycles_make_high and bool(has_cycles):
            tier = "HIGH"
        elif (treat_toprisk_as_medium and top_risk_cells > 0) or (treat_volatile_as_medium and vol > 0) or (treat_symbolic_as_medium and sym > 0) or (treat_constants_as_medium and consts > 0) or (treat_orphans_as_medium and orphans > 0):
            tier = "MEDIUM"
        else:
            tier = "LOW"

        reason_bits: List[str] = []
        if has_cycles:
            reason_bits.append("circular dependency")
        if top_risk_cells > 0:
            reason_bits.append(f"{top_risk_cells} high-impact cell(s)")
        if vol > 0:
            reason_bits.append(f"{vol} volatile hit(s)")
        if sym > 0:
            reason_bits.append(f"{sym} symbolic ref(s)")
        if consts > 0:
            reason_bits.append(f"{consts} embedded constant(s)")
        if orphans > 0:
            reason_bits.append(f"{orphans} orphan formula(s)")
        reason = "; ".join(reason_bits)
        return tier, reason

    sheets_out: List[Dict[str, Any]] = []
    for sheet, count in per_sheet_formula_counts.items():
        sig = per_sheet_signals.get(sheet, {"volatile": 0, "symbolic": 0}) or {}
        vol = int(sig.get("volatile", 0) or 0)
        sym = int(sig.get("symbolic", 0) or 0)
        consts = int(sig.get("constants", 0) or 0)
        circ_cells = len(circular_cells_by_sheet.get(sheet, set()) or set())
        top_cnt = int(top_by_sheet.get(sheet, 0) or 0)

        risk, reason = _compute_risk_tier(
            scope="sheet",
            signals_map={"volatile": vol, "symbolic": sym, "constants": consts, "orphans": int(sig.get("orphans", 0) or 0)},
            has_cycles=(circ_cells > 0),
            top_risk_cells=top_cnt,
            scoring_cfg=scoring,
        )

        sheets_out.append(
            {
                "sheet": sheet,
                "risk": risk,
                "formulas": int(count),
                "circular_cells": int(circ_cells),
                "high_risk_cells": int(top_cnt),
                "reason": reason,
            }
        )


# workbook risk summary
    reason_bits: List[str] = []
    if len(cycles) > 0:
        reason_bits.append(f"{len(cycles)} circular reference cycle(s)")
    if len(top_risks) > 0:
        reason_bits.append(f"{len(top_risks)} high-impact cells in Top-N list")
    if volatile_hits > 0:
        reason_bits.append(f"{volatile_hits} volatile function hit(s)")
    if symbolic_refs > 0:
        reason_bits.append(f"{symbolic_refs} symbolic reference(s)")
    if hardcoded_constants_hits > 0:
        reason_bits.append(f"{hardcoded_constants_hits} embedded constant(s)")
    if orphan_hits > 0:
        reason_bits.append(f"{orphan_hits} orphan formula(s)")

    # workbook risk tier (scoring if configured, else conservative fallback)
    workbook_risk, wb_reason = _compute_risk_tier(
        scope="workbook",
        signals_map={
            "volatile": int(volatile_hits),
            "symbolic": int(symbolic_refs),
            "constants": int(hardcoded_constants_hits),
        },
        has_cycles=(len(cycles) > 0),
        top_risk_cells=int(len(top_risks)),
        scoring_cfg=scoring,
    )

    # Reason shown in report: scored reason if available, else full signal list
    reason = wb_reason or ("; ".join(reason_bits) if reason_bits else "")

    scan: Dict[str, Any] = {
        "workbook": p.name,
        "generated": _utc_iso(),
        "trial": bool(trial_info["enabled"]),
        "trial_info": {
            "enabled": bool(trial_info["enabled"]),
            "formula_limit": int(trial_info["formula_limit"]),
            "formulas_scanned": int(formula_scanned),
            "truncated": bool(truncated),
        },
        "workbook_risk": workbook_risk,
        "reason": reason,
        "signals": {
            "circular": int(len(cycles)),
            "volatile": int(volatile_hits),
            "symbolic": int(symbolic_refs),
            "hardcoded_constants": int(hardcoded_constants_hits),
            "orphan_formulas": int(orphan_hits),
            "drift_formulas": int(drift_hits),
            "top_risks": int(len(top_risks)),
        },
        "hardcoded_constants": hardcoded_constants_findings,
        "orphan_formulas": orphan_findings,
        "formula_drift": drift_findings,
        "sheets": sheets_out,
        "top_risks": top_risks,
        "cycles": cycles,
        "symbolic_edges": symbolic_edges,
        "errors": errors,
    }

    # Write artifacts
    json_path = outp / f"{p.stem}.sheetguard.step5.json"
    html_path = outp / f"{p.stem}.sheetguard.report.html"

    json_path.write_text(json.dumps(scan, indent=2, ensure_ascii=False), encoding="utf-8")
    html = render_html_report(scan)
    html_path.write_text(html, encoding="utf-8")

    return scan, str(json_path), str(html_path)

def run_scan(
    workbook_path: str,
    out_dir: str,
    top_n: int = 10,
    trial: bool = False,
    open_report: bool = False,  # IMPORTANT: default False so UI won't double-open
) -> Tuple[str, str]:
    """
    UI-friendly wrapper:
      returns (json_path, html_path)
    """
    _scan, json_path, html_path = scan_workbook(
        workbook_path=workbook_path,
        out_dir=out_dir,
        top_n=top_n,
        trial=trial,
    )
    if open_report and html_path and os.path.exists(html_path):
        webbrowser.open_new_tab(Path(html_path).as_uri())
    return json_path, html_path


def main(argv: Optional[List[str]] = None) -> int:
    argv = list(sys.argv[1:] if argv is None else argv)

    parser = argparse.ArgumentParser(prog="sheetguard", description="SheetGuard - offline Excel diagnostic report")
    parser.add_argument("workbook", nargs="?", help="Path to .xlsx/.xlsm workbook")
    parser.add_argument("-o", "--out", dest="out_dir", default="", help="Output base directory (default: workbook folder)")
    parser.add_argument("--top", dest="top_n", type=int, default=10, help="Top-N high impact cells to list")
    parser.add_argument("--trial", action="store_true", help="Enable trial mode (formula limit from config)")
    parser.add_argument("--no-open", action="store_true", help="Do not open report automatically")

    args = parser.parse_args(argv)

    if not args.workbook:
        print("ERROR: No workbook provided.")
        return 2

    try:
        _scan, json_path, html_path = scan_workbook(
            workbook_path=args.workbook,
            out_dir=args.out_dir,
            top_n=int(args.top_n),
            trial=bool(args.trial),
        )
        print(f"JSON: {json_path}")
        print(f"HTML: {html_path}")
        if not args.no_open and html_path and os.path.exists(html_path):
            webbrowser.open_new_tab(Path(html_path).as_uri())
        return 0
    except InvalidFileException:
        print("ERROR: Invalid or unsupported workbook file.")
        return 3
    except Exception as e:
        print("ERROR:", e)
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
